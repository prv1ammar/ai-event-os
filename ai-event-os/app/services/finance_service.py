"""
app/services/finance_service.py
────────────────────────────────
Financial dashboard, KPI computation, and export generation
(Excel via openpyxl, PDF via ReportLab).

All monetary values are in MAD (Moroccan Dirham).

KPI formulas
────────────
ROI          = (revenue_confirmed - budget_total) / budget_total × 100
occupancy    = booked_stands / total_stands × 100
result_pct   = result_forecast / revenue_confirmed × 100
forecast     = expenses_committed × 1.10   (10 % contingency, capped at budget)
"""

from __future__ import annotations

import io
import uuid

from fastapi import HTTPException, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.booth import Booth
from app.models.budget import BudgetCategory, Expense
from app.models.event import Event
from app.models.payment import Payment


# ── Internal helper ────────────────────────────────────────────────────────────

async def _get_event_or_404(db: AsyncSession, event_id: uuid.UUID) -> Event:
    result = await db.execute(select(Event).where(Event.id == event_id))
    event = result.scalar_one_or_none()
    if event is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Event {event_id} not found",
        )
    return event


# ── Core dashboard ─────────────────────────────────────────────────────────────

async def get_financial_dashboard(db: AsyncSession, event_id: uuid.UUID) -> dict:
    """
    Returns the complete financial dashboard dictionary with all KPIs,
    revenue breakdown and budget-vs-actual per category.
    """
    event = await _get_event_or_404(db, event_id)

    # ── Revenue ───────────────────────────────────────────────────────────────
    rev_q = await db.execute(
        select(func.coalesce(func.sum(Payment.amount_mad), 0))
        .where(Payment.event_id == event_id, Payment.status == "paid")
    )
    revenue_confirmed_mad: int = int(rev_q.scalar() or 0)

    # Revenue target comes from the event's budget field (used as revenue target)
    revenue_target_mad: int = int(event.budget or 0)

    # ── Budget / Expenses ─────────────────────────────────────────────────────
    budget_q = await db.execute(
        select(func.coalesce(func.sum(BudgetCategory.budget_mad), 0))
        .where(BudgetCategory.event_id == event_id)
    )
    budget_total_mad: int = int(budget_q.scalar() or 0)

    # All non-cancelled expenses = "committed"
    committed_q = await db.execute(
        select(func.coalesce(func.sum(Expense.amount_mad), 0))
        .where(Expense.event_id == event_id, Expense.status != "cancelled")
    )
    expenses_committed_mad: int = int(committed_q.scalar() or 0)

    # Only expenses with status = paid
    paid_q = await db.execute(
        select(func.coalesce(func.sum(Expense.amount_mad), 0))
        .where(Expense.event_id == event_id, Expense.status == "paid")
    )
    expenses_paid_mad: int = int(paid_q.scalar() or 0)

    remaining_to_spend_mad: int = budget_total_mad - expenses_committed_mad
    raw_forecast = int(expenses_committed_mad * 1.10)
    forecast_final_mad: int = (
        min(raw_forecast, budget_total_mad) if budget_total_mad > 0 else raw_forecast
    )

    # ── Financial result ──────────────────────────────────────────────────────
    result_forecast_mad: int = revenue_confirmed_mad - budget_total_mad
    result_percentage: float = (
        round(result_forecast_mad / revenue_confirmed_mad * 100, 1)
        if revenue_confirmed_mad > 0
        else 0.0
    )

    # ── ROI ────────────────────────────────────────────────────────────────────
    roi_percent: float = (
        round((revenue_confirmed_mad - budget_total_mad) / budget_total_mad * 100, 1)
        if budget_total_mad > 0
        else 0.0
    )

    # ── Occupancy rate ─────────────────────────────────────────────────────────
    total_booths_q = await db.execute(
        select(func.count(Booth.id)).where(Booth.event_id == event_id)
    )
    total_booths: int = int(total_booths_q.scalar() or 0)

    booked_q = await db.execute(
        select(func.count(Booth.id))
        .where(Booth.event_id == event_id, Booth.status != "available")
    )
    booked_booths: int = int(booked_q.scalar() or 0)

    occupancy_rate: float = (
        round(booked_booths / total_booths * 100, 1) if total_booths > 0 else 0.0
    )

    # ── Revenue by source ──────────────────────────────────────────────────────
    source_q = await db.execute(
        select(
            Payment.source,
            func.coalesce(func.sum(Payment.amount_mad), 0).label("total"),
            func.count(Payment.id).label("count"),
        )
        .where(Payment.event_id == event_id, Payment.status == "paid")
        .group_by(Payment.source)
    )
    revenue_by_source: list[dict] = []
    for row in source_q:
        amount = int(row.total)
        pct = (
            round(amount / revenue_confirmed_mad * 100, 1)
            if revenue_confirmed_mad > 0
            else 0.0
        )
        revenue_by_source.append(
            {
                "source": row.source or "other",
                "amount_mad": amount,
                "count": int(row.count),
                "percentage": pct,
            }
        )

    # ── Budget by category with variance ──────────────────────────────────────
    cats_q = await db.execute(
        select(BudgetCategory).where(BudgetCategory.event_id == event_id)
    )
    categories = list(cats_q.scalars().all())

    spent_q = await db.execute(
        select(
            Expense.category_id,
            func.coalesce(func.sum(Expense.amount_mad), 0).label("spent"),
        )
        .where(Expense.event_id == event_id, Expense.status != "cancelled")
        .group_by(Expense.category_id)
    )
    spent_map: dict[str, int] = {
        str(row.category_id): int(row.spent) for row in spent_q
    }

    budget_by_category: list[dict] = []
    for cat in categories:
        spent = spent_map.get(str(cat.id), 0)
        variance = cat.budget_mad - spent
        budget_by_category.append(
            {
                "category": cat.name,
                "budget_mad": cat.budget_mad,
                "spent_mad": spent,
                "variance_mad": variance,
                "variance_pct": (
                    round(variance / cat.budget_mad * 100, 1)
                    if cat.budget_mad > 0
                    else 0.0
                ),
            }
        )

    return {
        "event_id": str(event_id),
        "event_name": event.name,
        # ── Budget ─────────────────────────────────────────────────────────────
        "budget_total_mad": budget_total_mad,
        "expenses_committed_mad": expenses_committed_mad,
        "expenses_paid_mad": expenses_paid_mad,
        "remaining_to_spend_mad": remaining_to_spend_mad,
        "forecast_final_mad": forecast_final_mad,
        # ── Revenue ────────────────────────────────────────────────────────────
        "revenue_confirmed_mad": revenue_confirmed_mad,
        "revenue_target_mad": revenue_target_mad,
        "result_forecast_mad": result_forecast_mad,
        "result_percentage": result_percentage,
        # ── KPIs ───────────────────────────────────────────────────────────────
        "roi_percent": roi_percent,
        "occupancy_rate": occupancy_rate,
        # ── Breakdowns ─────────────────────────────────────────────────────────
        "revenue_by_source": revenue_by_source,
        "budget_by_category": budget_by_category,
    }


async def get_kpis(db: AsyncSession, event_id: uuid.UUID) -> dict:
    """Lightweight KPI card data — subset of the full dashboard."""
    d = await get_financial_dashboard(db, event_id)
    return {
        "event_id": d["event_id"],
        "event_name": d["event_name"],
        "revenue_confirmed_mad": d["revenue_confirmed_mad"],
        "budget_total_mad": d["budget_total_mad"],
        "expenses_committed_mad": d["expenses_committed_mad"],
        "result_forecast_mad": d["result_forecast_mad"],
        "roi_percent": d["roi_percent"],
        "occupancy_rate": d["occupancy_rate"],
    }


async def get_revenue_by_source(
    db: AsyncSession, event_id: uuid.UUID
) -> list[dict]:
    """Revenue breakdown by source for a given event."""
    d = await get_financial_dashboard(db, event_id)
    return d["revenue_by_source"]


# ── Excel export ───────────────────────────────────────────────────────────────

def export_financial_excel(event_name: str, financial_data: dict) -> bytes:
    """
    4-sheet Excel workbook:
      Sheet 1 — Résumé        : KPI cards
      Sheet 2 — Budget        : budget vs. actual per category
      Sheet 3 — Paiements     : payment register header (live data via API)
      Sheet 4 — Revenus/source: revenue breakdown table

    Currency format : '# ##0 "MAD"'
    Row colours     : paid → green, pending → yellow, negative variance → red
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # ── Palette ────────────────────────────────────────────────────────────────
    NAVY = "1A1A2E"
    WHITE = "FFFFFF"
    GREY = "F3F4F6"
    GREEN = "DCFCE7"
    RED = "FEE2E2"
    YELLOW = "FEF9C3"

    def header(cell, text: str) -> None:
        cell.value = text
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def mad(val: int) -> str:
        return f"{val:,} MAD".replace(",", " ")

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 1 — Résumé
    # ──────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 28

    header(ws1["A1"], "Indicateur")
    header(ws1["B1"], "Valeur")
    ws1.row_dimensions[1].height = 22

    kpis = [
        ("Événement", event_name),
        ("CA confirmé (MAD)", mad(financial_data.get("revenue_confirmed_mad", 0))),
        ("Objectif revenus (MAD)", mad(financial_data.get("revenue_target_mad", 0))),
        ("Budget total (MAD)", mad(financial_data.get("budget_total_mad", 0))),
        ("Dépenses engagées (MAD)", mad(financial_data.get("expenses_committed_mad", 0))),
        ("Dépenses payées (MAD)", mad(financial_data.get("expenses_paid_mad", 0))),
        ("Reste à dépenser (MAD)", mad(financial_data.get("remaining_to_spend_mad", 0))),
        ("Prévision finale (MAD)", mad(financial_data.get("forecast_final_mad", 0))),
        ("Résultat prévisionnel (MAD)", mad(financial_data.get("result_forecast_mad", 0))),
        ("ROI (%)", f"{financial_data.get('roi_percent', 0.0):.1f} %"),
        ("Taux d'occupation stands (%)", f"{financial_data.get('occupancy_rate', 0.0):.1f} %"),
    ]
    for i, (label, value) in enumerate(kpis, start=2):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = value
        if i % 2 == 0:
            for col in ["A", "B"]:
                ws1[f"{col}{i}"].fill = PatternFill("solid", fgColor=GREY)

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 2 — Budget
    # ──────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Budget")
    for col, width in zip(["A", "B", "C", "D", "E"], [22, 20, 20, 20, 15]):
        ws2.column_dimensions[col].width = width

    for col, title in enumerate(
        ["Catégorie", "Budget (MAD)", "Dépensé (MAD)", "Variance (MAD)", "Variance (%)"],
        start=1,
    ):
        header(ws2.cell(row=1, column=col), title)

    for r, item in enumerate(financial_data.get("budget_by_category", []), start=2):
        ws2.cell(row=r, column=1, value=item["category"])
        ws2.cell(row=r, column=2, value=item["budget_mad"])
        ws2.cell(row=r, column=3, value=item["spent_mad"])
        var_cell = ws2.cell(row=r, column=4, value=item["variance_mad"])
        ws2.cell(row=r, column=5, value=f"{item.get('variance_pct', 0):.1f} %")
        var_cell.fill = PatternFill(
            "solid", fgColor=GREEN if item["variance_mad"] >= 0 else RED
        )

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 3 — Paiements  (header skeleton; live data from /api/v1/payments)
    # ──────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Paiements")
    for col, (title, width) in enumerate(
        zip(
            ["ID", "Montant (MAD)", "Méthode", "Source", "Statut", "Date paiement"],
            [38, 20, 16, 18, 14, 22],
        ),
        start=1,
    ):
        ws3.column_dimensions[chr(64 + col)].width = width
        header(ws3.cell(row=1, column=col), title)

    ws3["A2"] = "Données en temps réel — voir GET /api/v1/payments?event_id=…"
    ws3["A2"].font = Font(italic=True, color="6B7280")

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 4 — Revenus par source
    # ──────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Revenus par source")
    for col, (title, width) in enumerate(
        zip(["Source", "Montant (MAD)", "Nb. paiements", "Part (%)"], [22, 22, 18, 14]),
        start=1,
    ):
        ws4.column_dimensions[chr(64 + col)].width = width
        header(ws4.cell(row=1, column=col), title)

    for r, item in enumerate(financial_data.get("revenue_by_source", []), start=2):
        ws4.cell(row=r, column=1, value=item["source"])
        ws4.cell(row=r, column=2, value=item["amount_mad"])
        ws4.cell(row=r, column=3, value=item.get("count", 0))
        ws4.cell(row=r, column=4, value=f"{item.get('percentage', 0):.1f} %")
        if r % 2 == 0:
            for c in range(1, 5):
                ws4.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREY)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF summary export ─────────────────────────────────────────────────────────

def export_financial_pdf(event_name: str, financial_data: dict) -> bytes:
    """
    A4 financial summary PDF with:
    - KPI table
    - Budget-vs-actual table
    - Revenue-by-source table
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
    )

    PRIMARY = colors.HexColor("#1A1A2E")
    ACCENT = colors.HexColor("#7C3AED")
    LIGHT = colors.HexColor("#F3F4F6")
    GREEN = colors.HexColor("#DCFCE7")
    RED = colors.HexColor("#FEE2E2")
    WHITE = colors.white

    styles = getSampleStyleSheet()

    def ps(name: str, **kw) -> ParagraphStyle:
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    def mad(val: int) -> str:
        return f"{val:,} MAD".replace(",", " ")

    title_s = ps("T", fontSize=20, textColor=PRIMARY, fontName="Helvetica-Bold", spaceAfter=3 * mm)
    sub_s = ps("S", fontSize=10, textColor=ACCENT, spaceAfter=5 * mm)
    sec_s = ps("Sec", fontSize=13, textColor=PRIMARY, fontName="Helvetica-Bold",
               spaceBefore=5 * mm, spaceAfter=3 * mm)
    foot_s = ps("F", fontSize=7, textColor=colors.grey)

    tbl_header = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ])

    elems = []

    # Header
    elems.append(Paragraph("AI EVENT OS", title_s))
    elems.append(Paragraph(f"Rapport Financier — {event_name}", sub_s))
    elems.append(HRFlowable(width="100%", thickness=2, color=ACCENT))
    elems.append(Spacer(1, 5 * mm))

    # KPI table
    elems.append(Paragraph("Indicateurs Clés de Performance", sec_s))
    kpi_rows = [
        ["Indicateur", "Valeur"],
        ["CA confirmé (MAD)", mad(financial_data.get("revenue_confirmed_mad", 0))],
        ["Budget total (MAD)", mad(financial_data.get("budget_total_mad", 0))],
        ["Dépenses engagées (MAD)", mad(financial_data.get("expenses_committed_mad", 0))],
        ["Résultat prévisionnel (MAD)", mad(financial_data.get("result_forecast_mad", 0))],
        ["ROI (%)", f"{financial_data.get('roi_percent', 0):.1f} %"],
        ["Taux d'occupation (%)", f"{financial_data.get('occupancy_rate', 0):.1f} %"],
    ]
    t = Table(kpi_rows, colWidths=[105 * mm, 70 * mm])
    t.setStyle(tbl_header)
    elems.extend([t, Spacer(1, 5 * mm)])

    # Budget table
    elems.append(Paragraph("Budget par Catégorie", sec_s))
    bud_rows = [["Catégorie", "Budget (MAD)", "Dépensé (MAD)", "Variance (MAD)"]]
    for item in financial_data.get("budget_by_category", []):
        bud_rows.append([
            item["category"],
            mad(item["budget_mad"]),
            mad(item["spent_mad"]),
            mad(item["variance_mad"]),
        ])
    if len(bud_rows) > 1:
        bt = Table(bud_rows, colWidths=[44 * mm, 44 * mm, 44 * mm, 44 * mm])
        bt_style = TableStyle(list(tbl_header._cmds))
        for i, item in enumerate(financial_data.get("budget_by_category", []), start=1):
            clr = GREEN if item["variance_mad"] >= 0 else RED
            bt_style.add("BACKGROUND", (3, i), (3, i), clr)
        bt.setStyle(bt_style)
        elems.extend([bt, Spacer(1, 5 * mm)])
    else:
        elems.append(Paragraph("Aucune catégorie définie.", styles["Normal"]))

    # Revenue by source
    elems.append(Paragraph("Revenus par Source", sec_s))
    rev_rows = [["Source", "Montant (MAD)", "Part (%)"]]
    for item in financial_data.get("revenue_by_source", []):
        rev_rows.append([item["source"], mad(item["amount_mad"]),
                         f"{item.get('percentage', 0):.1f} %"])
    if len(rev_rows) > 1:
        rt = Table(rev_rows, colWidths=[60 * mm, 60 * mm, 56 * mm])
        rt.setStyle(tbl_header)
        elems.extend([rt, Spacer(1, 5 * mm)])
    else:
        elems.append(Paragraph("Aucun paiement confirmé.", styles["Normal"]))

    # Footer
    elems.append(Spacer(1, 8 * mm))
    elems.append(HRFlowable(width="100%", thickness=0.5, color=ACCENT))
    elems.append(Spacer(1, 2 * mm))
    elems.append(Paragraph(
        "Généré par AI EVENT OS — Tous montants en MAD (Dirham Marocain) — TVA 20 % applicable",
        foot_s,
    ))

    doc.build(elems)
    return buf.getvalue()
