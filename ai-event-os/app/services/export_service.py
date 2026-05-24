"""
app/services/export_service.py
────────────────────────────────
Excel export for leads using openpyxl.

Workbook layout:
  Sheet 1 — "Leads"         all leads, color-coded by status
  Sheet 2 — "Statistiques"  count / % / avg score per status
  Sheet 3 — "Top Leads"     only leads with score >= 70, sorted desc
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Status → background hex colour ────────────────────────────────────────────

STATUS_FILL: dict[str, str] = {
    "new":          "FFFFFF",   # white
    "contacted":    "DBEAFE",   # light blue
    "qualified":    "DCFCE7",   # light green
    "opportunity":  "FEF9C3",   # light gold
    "closed_won":   "BBF7D0",   # green
    "closed_lost":  "FEE2E2",   # light red
}

HEADER_BG   = "1A1A2E"  # dark navy
HEADER_FG   = "FFFFFF"

COLUMNS = [
    ("Nom",              "visitor_name"),
    ("Email",            "visitor_email"),
    ("Téléphone",        "visitor_phone"),
    ("Entreprise",       "visitor_company"),
    ("Exposant",         "exhibitor_name"),
    ("Statut",           "status"),
    ("Score",            "score"),
    ("Budget (MAD)",     "budget_range"),
    ("Notes",            "notes"),
    ("Date création",    "created_at"),
]


def _header_row(ws, headers: list[str]) -> None:
    """Write a styled header row to *ws*."""
    h_font  = Font(bold=True, color=HEADER_FG)
    h_fill  = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center")

    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = h_font
        cell.fill  = h_fill
        cell.alignment = h_align


def _auto_width(ws) -> None:
    """Adjust each column width to the longest value (capped at 50)."""
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def export_leads_excel(leads_data: list[dict[str, Any]]) -> bytes:
    """
    Convert a list of lead dicts into a formatted Excel workbook.

    Each dict in *leads_data* must include the keys defined in COLUMNS above.
    Returns raw bytes ready to be streamed as a Response.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: All leads ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Leads"
    ws1.freeze_panes = "A2"   # keep header visible when scrolling

    col_labels = [c[0] for c in COLUMNS]
    col_keys   = [c[1] for c in COLUMNS]
    _header_row(ws1, col_labels)

    for row_idx, lead in enumerate(leads_data, 2):
        st    = lead.get("status", "new")
        color = STATUS_FILL.get(st, "FFFFFF")
        fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for col_idx, key in enumerate(col_keys, 1):
            val  = lead.get(key)
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    _auto_width(ws1)

    # ── Sheet 2: Summary statistics ────────────────────────────────────────────
    ws2 = wb.create_sheet("Statistiques")
    _header_row(ws2, ["Statut", "Nombre", "% du Total", "Score Moyen"])

    # Aggregate
    status_counts: dict[str, int]       = {}
    status_scores: dict[str, list[int]] = {}

    for lead in leads_data:
        s  = lead.get("status", "new")
        sc = int(lead.get("score") or 0)
        status_counts[s] = status_counts.get(s, 0) + 1
        status_scores.setdefault(s, []).append(sc)

    total = len(leads_data)
    for s in sorted(status_counts):
        cnt      = status_counts[s]
        pct      = round(cnt / total * 100, 1) if total else 0.0
        avg_sc   = round(sum(status_scores[s]) / len(status_scores[s]), 1)
        color    = STATUS_FILL.get(s, "FFFFFF")
        row_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        r = ws2.max_row + 1
        for ci, v in enumerate([s, cnt, f"{pct}%", avg_sc], 1):
            cell = ws2.cell(row=r, column=ci, value=v)
            cell.fill = row_fill

    _auto_width(ws2)

    # ── Sheet 3: Top leads (score >= 70) ───────────────────────────────────────
    ws3 = wb.create_sheet("Top Leads")
    _header_row(ws3, col_labels)

    gold_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    top = sorted(
        [l for l in leads_data if int(l.get("score") or 0) >= 70],
        key=lambda x: int(x.get("score") or 0),
        reverse=True,
    )

    for row_idx, lead in enumerate(top, 2):
        for col_idx, key in enumerate(col_keys, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=lead.get(key))
            cell.fill = gold_fill

    _auto_width(ws3)

    # ── Serialise ──────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
